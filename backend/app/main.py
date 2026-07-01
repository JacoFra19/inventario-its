from datetime import datetime, timedelta
from fastapi import FastAPI, HTTPException, Body, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, or_
from io import BytesIO
import hashlib
import qrcode

from .db import engine, SessionLocal
from .models import Base, Location, LocationCounter, Category, Item, Asset, Assignee, AssetMovement, AssetLog, StockCard, StockMovement, Event, EventAsset, EventStock
from .seed import seed_categories, seed_locations

app = FastAPI(title="Inventario ITS", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Solo ambiente di sviluppo
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

Base.metadata.create_all(bind=engine)
seed_locations()
seed_categories()



def generate_inventory_code(db: Session, location_code: str) -> str:
    # Trova la location
    loc = db.execute(select(Location).where(Location.code == location_code)).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail=f"Sede non trovata: {location_code}")

    # Prendi o crea counter della sede
    counter = db.execute(select(LocationCounter).where(LocationCounter.location_id == loc.id)).scalar_one_or_none()
    if not counter:
        counter = LocationCounter(location_id=loc.id, next_number=1)
        db.add(counter)
        db.flush()  # assegna

    n = counter.next_number
    if n > 9999:
        raise HTTPException(status_code=400, detail=f"Progressivo esaurito per sede {location_code}")

    code = f"ITST-{loc.code}-{n:04d}"

    # incrementa per il prossimo
    counter.next_number = n + 1
    return code


# --- HELPER FOR ITEM SERIALIZATION ---
def serialize_item(
    item: Item,
    asset_count: int | None = None,
    stock_card_count: int | None = None,
):
    category = item.category

    return {
        "id": item.id,
        "name": item.name,
        "category_id": item.category_id,
        "category": None if not category else {
            "id": category.id,
            "name": category.name,
        },
        "brand": item.brand,
        "model": item.model,
        "technical_specs": item.technical_specs,
        "is_serialized": item.is_serialized,
        "asset_count": asset_count,
        "stock_card_count": stock_card_count,
    }


def serialize_assignee(
    assignee: Assignee,
    asset_count: int | None = None,
    assets: list[Asset] | None = None,
):
    result = {
        "id": assignee.id,
        "name": assignee.name,
        "type": assignee.type,
        "email": assignee.email,
        "phone": assignee.phone,
        "notes": assignee.notes,
        "is_active": assignee.is_active,
        "created_at": assignee.created_at,
        "asset_count": asset_count,
    }

    if assets is not None:
        result["assets"] = [
            {
                "id": asset.id,
                "inventory_code": asset.inventory_code,
                "status": asset.status,
                "assigned_to": asset.assigned_to,
                "notes": asset.notes,
            }
            for asset in assets
        ]

    return result


def validate_assignee_type(assignee_type: str):
    normalized_type = assignee_type.strip().upper()
    allowed_types = {"PERSON", "DEPARTMENT", "OTHER"}

    if normalized_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail="Tipo assegnatario non valido. Usa PERSON, DEPARTMENT o OTHER.",
        )

    return normalized_type


# --- HELPER FOR ASSET LOGS ---
def create_asset_log(
    db: Session,
    asset_id: int,
    action_type: str,
    description: str,
    created_by: str | None = None,
):
    log = AssetLog(
        asset_id=asset_id,
        action_type=action_type,
        description=description,
        created_by=created_by,
    )
    db.add(log)
    return log


def location_export_label(location: Location | None):
    if not location:
        return ""

    return f"{location.code} - {location.name}"


def item_export_label(item: Item | None):
    if not item:
        return ""

    parts = [item.name]
    if item.brand:
        parts.append(item.brand)
    if item.model:
        parts.append(item.model)

    return " - ".join(parts)


def text_matches(query: str, *values: object):
    haystack = " ".join(
        str(value).lower()
        for value in values
        if value is not None and str(value).strip()
    )

    return query in haystack


def empty_search_response(query: str):
    return {
        "query": query,
        "results": {
            "assets": [],
            "items": [],
            "stocks": [],
            "events": [],
        },
    }


LOCATION_COORDINATES = {
    "BA1": {"lat": 41.1171, "lng": 16.8719},
    "BA2": {"lat": 41.1200, "lng": 16.8650},
    "BR1": {"lat": 40.6327, "lng": 17.9418},
    "CO1": {"lat": 40.9688, "lng": 17.1136},
    "FA1": {"lat": 40.8346, "lng": 17.3597},
    "GR1": {"lat": 40.8173, "lng": 16.4166},
    "LE1": {"lat": 40.3515, "lng": 18.1750},
    "LE2": {"lat": 40.3510, "lng": 18.1720},
    "MA1": {"lat": 40.3993, "lng": 17.6335},
    "MF1": {"lat": 41.6307, "lng": 15.9188},
    "MO1": {"lat": 39.8490, "lng": 18.3127},
    "PO1": {"lat": 40.0524, "lng": 18.3775},
    "PU1": {"lat": 40.8491, "lng": 17.1213},
    "TA1": {"lat": 40.4644, "lng": 17.2470},
    "TR1": {"lat": 41.2775, "lng": 16.4101},
}


IMPORT_SHEET_NAME = "Import"
IMPORT_MAX_ROWS = 1000
IMPORT_COLUMNS = [
    "tipo",
    "sede",
    "categoria",
    "nome",
    "marca",
    "modello",
    "serializzato",
    "quantita",
    "soglia_minima",
    "note",
    "assegnatario",
]


def normalize_import_text(value):
    if value is None:
        return ""

    return str(value).strip()


def normalize_import_key(value):
    return normalize_import_text(value).lower()


def parse_import_bool(value):
    if value is None or normalize_import_text(value) == "":
        return None

    normalized = normalize_import_text(value).lower()
    if normalized in {"si", "sì", "yes", "true", "1", "asset"}:
        return True
    if normalized in {"no", "false", "0", "stock"}:
        return False

    raise ValueError("Valore serializzato non valido. Usa SI/NO.")


def parse_import_int(value, field_name: str):
    if value is None or normalize_import_text(value) == "":
        return None

    try:
        parsed = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field_name} deve essere un numero intero.")

    return parsed


def import_result_template():
    return {
        "rows": [],
        "summary": {
            "total_rows": 0,
            "valid_rows": 0,
            "warning_rows": 0,
            "error_rows": 0,
            "items_to_create": 0,
            "items_to_reuse": 0,
            "categories_to_create": 0,
            "asset_to_create": 0,
            "stockcard_to_create": 0,
            "stockcard_to_update": 0,
            "assignee_to_create": 0,
            "assignee_to_reuse": 0,
        },
        "can_commit": False,
        "operations": [],
    }


def find_category_by_name(db: Session, name: str):
    normalized = normalize_import_key(name)
    for category in db.query(Category).all():
        if normalize_import_key(category.name) == normalized:
            return category

    return None


def find_location_by_code(db: Session, code: str):
    normalized = normalize_import_key(code)
    for location in db.query(Location).all():
        if normalize_import_key(location.code) == normalized:
            return location

    return None


def find_item(
    db: Session,
    category_name: str,
    name: str,
    brand: str | None,
    model: str | None,
    is_serialized: bool,
):
    category_key = normalize_import_key(category_name)
    name_key = normalize_import_key(name)
    brand_key = normalize_import_key(brand)
    model_key = normalize_import_key(model)

    for item in db.query(Item).all():
        category = item.category
        if (
            category
            and normalize_import_key(category.name) == category_key
            and normalize_import_key(item.name) == name_key
            and normalize_import_key(item.brand) == brand_key
            and normalize_import_key(item.model) == model_key
            and item.is_serialized == is_serialized
        ):
            return item

    return None


def find_assignee_by_name(db: Session, name: str):
    normalized = normalize_import_key(name)
    for assignee in db.query(Assignee).all():
        if normalize_import_key(assignee.name) == normalized:
            return assignee

    return None


def build_import_preview(file_bytes: bytes, db: Session):
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"File Excel non valido: {exc}")

    if IMPORT_SHEET_NAME not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Foglio '{IMPORT_SHEET_NAME}' non trovato.")

    sheet = workbook[IMPORT_SHEET_NAME]
    headers = [
        normalize_import_key(cell.value).replace(" ", "_")
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]

    missing_columns = [column for column in IMPORT_COLUMNS if column not in headers]
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Colonne mancanti nel template: {', '.join(missing_columns)}",
        )

    column_indexes = {header: index for index, header in enumerate(headers)}
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    non_empty_rows = [
        row for row in data_rows if any(normalize_import_text(value) for value in row)
    ]

    if len(non_empty_rows) > IMPORT_MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"Il file contiene più di {IMPORT_MAX_ROWS} righe importabili.",
        )

    result = import_result_template()
    planned_categories: set[str] = set()
    planned_items: set[tuple[str, str, str, str, bool]] = set()
    reused_items: set[tuple[str, str, str, str, bool]] = set()
    planned_assignees: set[str] = set()
    reused_assignees: set[str] = set()
    planned_stockcards: set[tuple[str, str, str, str, bool, str]] = set()
    updated_stockcards: set[tuple[str, str, str, str, bool, str]] = set()

    def value(row, column):
        index = column_indexes[column]
        return row[index] if index < len(row) else None

    for offset, row in enumerate(data_rows, start=2):
        if not any(normalize_import_text(cell) for cell in row):
            continue

        errors = []
        warnings = []

        tipo = normalize_import_text(value(row, "tipo")).upper()
        sede = normalize_import_text(value(row, "sede")).upper()
        categoria = normalize_import_text(value(row, "categoria"))
        nome = normalize_import_text(value(row, "nome"))
        marca = normalize_import_text(value(row, "marca"))
        modello = normalize_import_text(value(row, "modello"))
        note = normalize_import_text(value(row, "note"))
        assegnatario = normalize_import_text(value(row, "assegnatario"))

        try:
            serializzato = parse_import_bool(value(row, "serializzato"))
        except ValueError as exc:
            serializzato = None
            errors.append(str(exc))

        try:
            quantita = parse_import_int(value(row, "quantita"), "quantita")
        except ValueError as exc:
            quantita = None
            errors.append(str(exc))

        try:
            soglia_minima = parse_import_int(value(row, "soglia_minima"), "soglia_minima")
        except ValueError as exc:
            soglia_minima = None
            errors.append(str(exc))

        if tipo not in {"ASSET", "STOCK"}:
            errors.append("Tipo obbligatorio: usa ASSET o STOCK.")

        location = find_location_by_code(db, sede) if sede else None
        if not sede:
            errors.append("Sede obbligatoria.")
        elif not location:
            errors.append(f"Sede non valida: {sede}.")

        if not categoria:
            errors.append("Categoria obbligatoria.")

        if not nome:
            errors.append("Nome item obbligatorio.")

        if quantita is None:
            errors.append("Quantita obbligatoria.")
        elif quantita < 1:
            errors.append("Quantita deve essere almeno 1.")

        if tipo == "ASSET":
            if serializzato is False:
                errors.append("Le righe ASSET devono avere serializzato=SI.")
            serializzato = True
            if soglia_minima not in {None, 0}:
                warnings.append("Soglia minima ignorata per righe ASSET.")

        if tipo == "STOCK":
            if serializzato is True:
                errors.append("Le righe STOCK devono avere serializzato=NO.")
            serializzato = False
            if assegnatario:
                warnings.append("Assegnatario ignorato per righe STOCK.")
            if soglia_minima is not None and soglia_minima < 0:
                errors.append("Soglia minima non può essere negativa.")

        status = "ERROR" if errors else "WARNING" if warnings else "VALID"
        item_key = (
            normalize_import_key(categoria),
            normalize_import_key(nome),
            normalize_import_key(marca),
            normalize_import_key(modello),
            bool(serializzato),
        )

        description_parts = [tipo or "-", nome or "-", sede or "-"]
        description = " - ".join(description_parts)

        if not errors:
            category = find_category_by_name(db, categoria)
            if category is None:
                planned_categories.add(normalize_import_key(categoria))

            item = find_item(db, categoria, nome, marca, modello, bool(serializzato))
            if item is None:
                planned_items.add(item_key)
            else:
                reused_items.add(item_key)

            if tipo == "ASSET":
                result["summary"]["asset_to_create"] += quantita or 0
                if assegnatario:
                    assignee = find_assignee_by_name(db, assegnatario)
                    if assignee is None:
                        planned_assignees.add(normalize_import_key(assegnatario))
                    else:
                        reused_assignees.add(normalize_import_key(assegnatario))

            if tipo == "STOCK" and location:
                stock_key = (*item_key, normalize_import_key(location.code))
                existing_stock = None
                if item:
                    existing_stock = (
                        db.query(StockCard)
                        .filter(
                            StockCard.item_id == item.id,
                            StockCard.location_id == location.id,
                        )
                        .first()
                    )

                if existing_stock:
                    updated_stockcards.add(stock_key)
                    warnings.append("Stockcard esistente: la quantità verrà incrementata.")
                    status = "WARNING"
                else:
                    planned_stockcards.add(stock_key)

            result["operations"].append({
                "row_number": offset,
                "type": tipo,
                "location_code": location.code if location else sede,
                "category": categoria,
                "name": nome,
                "brand": marca,
                "model": modello,
                "is_serialized": bool(serializzato),
                "quantity": quantita,
                "min_threshold": soglia_minima if soglia_minima is not None else 0,
                "notes": note,
                "assignee": assegnatario,
            })

        message = "; ".join(errors or warnings or ["Riga valida."])
        result["rows"].append({
            "row_number": offset,
            "type": tipo or "-",
            "description": description,
            "status": status,
            "message": message,
        })

    result["summary"]["total_rows"] = len(result["rows"])
    result["summary"]["valid_rows"] = sum(1 for row in result["rows"] if row["status"] == "VALID")
    result["summary"]["warning_rows"] = sum(1 for row in result["rows"] if row["status"] == "WARNING")
    result["summary"]["error_rows"] = sum(1 for row in result["rows"] if row["status"] == "ERROR")
    result["summary"]["items_to_create"] = len(planned_items)
    result["summary"]["items_to_reuse"] = len(reused_items)
    result["summary"]["categories_to_create"] = len(planned_categories)
    result["summary"]["stockcard_to_create"] = len(planned_stockcards)
    result["summary"]["stockcard_to_update"] = len(updated_stockcards)
    result["summary"]["assignee_to_create"] = len(planned_assignees)
    result["summary"]["assignee_to_reuse"] = len(reused_assignees)
    result["can_commit"] = result["summary"]["total_rows"] > 0 and result["summary"]["error_rows"] == 0

    return result


def get_or_create_import_category(db: Session, name: str):
    category = find_category_by_name(db, name)
    if category:
        return category

    category = Category(name=name.strip())
    db.add(category)
    db.flush()
    return category


def get_or_create_import_item(
    db: Session,
    category: Category,
    name: str,
    brand: str,
    model: str,
    is_serialized: bool,
):
    item = find_item(db, category.name, name, brand, model, is_serialized)
    if item:
        return item

    item = Item(
        name=name.strip(),
        category_id=category.id,
        brand=brand.strip() if brand else None,
        model=model.strip() if model else None,
        technical_specs=None,
        is_serialized=is_serialized,
    )
    db.add(item)
    db.flush()
    item.category = category
    return item


def get_or_create_import_assignee(db: Session, name: str):
    assignee = find_assignee_by_name(db, name)
    if assignee:
        if not assignee.is_active:
            assignee.is_active = True
        return assignee

    assignee = Assignee(
        name=name.strip(),
        type="PERSON",
        is_active=True,
    )
    db.add(assignee)
    db.flush()
    return assignee


def commit_import_operations(plan, db: Session):
    created_assets = 0
    created_stockcards = 0
    updated_stockcards = 0
    created_items: set[int] = set()
    reused_items: set[int] = set()
    created_assignees: set[int] = set()
    reused_assignees: set[int] = set()

    for operation in plan["operations"]:
        location = find_location_by_code(db, operation["location_code"])
        if not location:
            raise HTTPException(
                status_code=400,
                detail=f"Sede non trovata durante il commit: {operation['location_code']}",
            )

        category_before = find_category_by_name(db, operation["category"])
        category = get_or_create_import_category(db, operation["category"])
        item_before = find_item(
            db,
            category.name,
            operation["name"],
            operation["brand"],
            operation["model"],
            operation["is_serialized"],
        )
        item = get_or_create_import_item(
            db,
            category,
            operation["name"],
            operation["brand"],
            operation["model"],
            operation["is_serialized"],
        )

        if item_before:
            reused_items.add(item.id)
        else:
            created_items.add(item.id)

        if operation["type"] == "ASSET":
            assignee = None
            if operation["assignee"]:
                assignee_before = find_assignee_by_name(db, operation["assignee"])
                assignee = get_or_create_import_assignee(db, operation["assignee"])
                if assignee_before:
                    reused_assignees.add(assignee.id)
                else:
                    created_assignees.add(assignee.id)

            for _ in range(operation["quantity"]):
                inventory_code = generate_inventory_code(db, location.code)
                asset = Asset(
                    inventory_code=inventory_code,
                    item_id=item.id,
                    current_location_id=location.id,
                    assignee_id=assignee.id if assignee else None,
                    assigned_to=assignee.name if assignee else None,
                    status="ASSEGNATO" if assignee else "IN_SEDE",
                    notes=operation["notes"] or None,
                )
                db.add(asset)
                db.flush()

                create_asset_log(
                    db,
                    asset.id,
                    "CREATE",
                    f"Asset creato da import Excel in sede {location.code} - {location.name}",
                )

                if assignee:
                    create_asset_log(
                        db,
                        asset.id,
                        "ASSIGN",
                        f"Asset assegnato a {assignee.name} da import Excel",
                    )

                created_assets += 1

        if operation["type"] == "STOCK":
            stock = (
                db.query(StockCard)
                .filter(
                    StockCard.item_id == item.id,
                    StockCard.location_id == location.id,
                )
                .first()
            )

            if stock:
                updated_stockcards += 1
                if operation["min_threshold"] is not None:
                    stock.min_threshold = operation["min_threshold"]
                if operation["notes"]:
                    stock.notes = operation["notes"]
            else:
                stock = StockCard(
                    item_id=item.id,
                    location_id=location.id,
                    quantity=0,
                    min_threshold=operation["min_threshold"] or 0,
                    notes=operation["notes"] or None,
                )
                db.add(stock)
                db.flush()
                created_stockcards += 1

            if operation["quantity"] > 0:
                stock.quantity += operation["quantity"]
                movement = StockMovement(
                    stock_card_id=stock.id,
                    movement_type="LOAD",
                    quantity=operation["quantity"],
                    notes="Carico iniziale da import Excel",
                )
                db.add(movement)

    return {
        "created_assets": created_assets,
        "created_stockcards": created_stockcards,
        "updated_stockcards": updated_stockcards,
        "created_items": len(created_items),
        "reused_items": len(reused_items),
        "created_assignees": len(created_assignees),
        "reused_assignees": len(reused_assignees),
    }


def style_export_sheet(sheet):
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="E5E7EB")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        adjusted_width = min(max(max_length + 2, 12), 48)
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width


def export_workbook(workbook, filename: str):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def append_activity(
    activities: list[dict],
    activity_type: str,
    title: str,
    description: str,
    timestamp: datetime | None,
    category: str,
    severity: str = "info",
    references: dict | None = None,
):
    if not timestamp:
        return

    references = references or {}
    href = references.get("href")
    activity_key = "|".join([
        activity_type,
        title,
        description,
        timestamp.isoformat(),
        category,
    ])

    activities.append({
        "id": f"{activity_type.lower()}-{hashlib.sha1(activity_key.encode('utf-8')).hexdigest()[:12]}",
        "type": activity_type,
        "title": title,
        "description": description,
        "timestamp": timestamp,
        "category": category,
        "severity": severity,
        "references": references,
        "href": href if isinstance(href, str) else None,
    })


def asset_log_category(log: AssetLog):
    description = (log.description or "").lower()

    if "import excel" in description:
        return "import"

    if log.action_type == "TRANSFER":
        return "transfer"

    if log.action_type in {"ASSIGN", "UNASSIGN"}:
        return "assignee"

    return "asset"


def asset_log_title(log: AssetLog, inventory_code: str):
    if log.action_type == "CREATE" and "import excel" in (log.description or "").lower():
        return "Asset importato"

    labels = {
        "CREATE": "Asset creato",
        "ASSIGN": "Asset assegnato",
        "UNASSIGN": "Assegnazione rimossa",
        "TRANSFER": "Asset trasferito",
        "MARK_MISSING": "Asset mancante",
        "RESTORE": "Asset ripristinato",
        "EVENT_OUT": "Asset uscito per evento",
        "EVENT_RETURN": "Asset rientrato da evento",
        "EVENT_MISSING": "Asset mancante in evento",
    }

    return labels.get(log.action_type, inventory_code)


def stock_activity_category(movement: StockMovement):
    notes = (movement.notes or "").lower()

    if "import excel" in notes:
        return "import"

    return "stock"


def normalize_activity_filter(value: str | None):
    if not value:
        return None

    normalized = value.strip().lower()
    aliases = {
        "asset": "asset",
        "stock": "stock",
        "event": "event",
        "events": "event",
        "eventi": "event",
        "import": "import",
        "trasferimenti": "transfer",
        "transfer": "transfer",
        "transfers": "transfer",
        "assegnazioni": "assignee",
        "assignee": "assignee",
        "assignees": "assignee",
        "sistema": "system",
        "system": "system",
    }

    return aliases.get(normalized, normalized)


def activity_search_blob(activity: dict):
    references = activity.get("references") or {}
    reference_values = " ".join(
        str(value) for value in references.values() if value is not None
    )

    return " ".join([
        str(activity.get("type") or ""),
        str(activity.get("title") or ""),
        str(activity.get("description") or ""),
        str(activity.get("category") or ""),
        str(activity.get("severity") or ""),
        reference_values,
    ]).lower()


def build_activity_log(db: Session, source_limit: int | None = None):
    activities = []

    assets = {asset.id: asset for asset in db.query(Asset).all()}
    items = {item.id: item for item in db.query(Item).all()}
    locations = {location.id: location for location in db.query(Location).all()}
    stocks = {stock.id: stock for stock in db.query(StockCard).all()}
    events = {event.id: event for event in db.query(Event).all()}

    def limited(query):
        return query.limit(source_limit) if source_limit else query

    asset_logs = limited(
        db.query(AssetLog).order_by(AssetLog.created_at.desc())
    ).all()

    for log in asset_logs:
        asset = assets.get(log.asset_id)
        inventory_code = asset.inventory_code if asset else f"Asset ID {log.asset_id}"

        severity = "success"
        if log.action_type in {"MARK_MISSING", "EVENT_MISSING"}:
            severity = "critical"

        append_activity(
            activities,
            log.action_type,
            asset_log_title(log, inventory_code),
            f"{inventory_code}: {log.description}",
            log.created_at,
            asset_log_category(log),
            severity,
            {
                "asset_log_id": log.id,
                "asset_id": log.asset_id,
                "inventory_code": asset.inventory_code if asset else None,
                "href": f"/assets/{asset.inventory_code}" if asset else None,
            },
        )

    asset_movements = limited(
        db.query(AssetMovement).order_by(AssetMovement.moved_at.desc())
    ).all()

    for movement in asset_movements:
        asset = assets.get(movement.asset_id)
        inventory_code = asset.inventory_code if asset else f"Asset ID {movement.asset_id}"
        from_location = location_export_label(locations.get(movement.from_location_id)) or "Sede non registrata"
        to_location = location_export_label(locations.get(movement.to_location_id)) or "Sede non registrata"

        append_activity(
            activities,
            "ASSET_TRANSFER",
            "Asset trasferito",
            f"{inventory_code}: {from_location} -> {to_location}",
            movement.moved_at,
            "transfer",
            "success",
            {
                "asset_movement_id": movement.id,
                "asset_id": movement.asset_id,
                "inventory_code": asset.inventory_code if asset else None,
                "href": f"/assets/{asset.inventory_code}" if asset else None,
            },
        )

    stock_movements = limited(
        db.query(StockMovement).order_by(StockMovement.created_at.desc())
    ).all()

    stock_labels = {
        "LOAD": "Carico stock",
        "UNLOAD": "Scarico stock",
        "RETURN": "Rientro stock",
        "ADJUST": "Correzione stock",
    }

    for movement in stock_movements:
        stock = stocks.get(movement.stock_card_id)
        item = items.get(stock.item_id) if stock else None
        location = locations.get(stock.location_id) if stock else None
        label = item_export_label(item) or f"Stockcard ID {movement.stock_card_id}"
        location_label = location_export_label(location)
        category = stock_activity_category(movement)

        append_activity(
            activities,
            f"STOCK_{movement.movement_type}",
            "Stock importato" if category == "import" else stock_labels.get(movement.movement_type, movement.movement_type),
            f"{label} {f'({location_label})' if location_label else ''} - {movement.quantity} pezzi",
            movement.created_at,
            category,
            "success",
            {
                "stock_movement_id": movement.id,
                "stock_card_id": movement.stock_card_id,
                "href": "/stocks",
            },
        )

    event_assets = limited(
        db.query(EventAsset).order_by(EventAsset.created_at.desc())
    ).all()

    for event_asset in event_assets:
        event = events.get(event_asset.event_id)
        asset = assets.get(event_asset.asset_id)
        event_name = event.name if event else f"Evento ID {event_asset.event_id}"
        inventory_code = asset.inventory_code if asset else f"Asset ID {event_asset.asset_id}"

        append_activity(
            activities,
            "EVENT_ASSET_OUT",
            "Asset aggiunto a evento",
            f"{inventory_code} -> {event_name}",
            event_asset.created_at,
            "event",
            "success",
            {
                "event_asset_id": event_asset.id,
                "event_id": event_asset.event_id,
                "asset_id": event_asset.asset_id,
                "inventory_code": asset.inventory_code if asset else None,
                "href": f"/events?eventId={event_asset.event_id}",
            },
        )

        if event_asset.returned_at:
            append_activity(
                activities,
                "EVENT_ASSET_RETURN",
                "Asset rientrato da evento",
                f"{inventory_code} <- {event_name}",
                event_asset.returned_at,
                "event",
                "success",
                {
                    "event_asset_id": event_asset.id,
                    "event_id": event_asset.event_id,
                    "asset_id": event_asset.asset_id,
                    "inventory_code": asset.inventory_code if asset else None,
                    "href": f"/events?eventId={event_asset.event_id}",
                },
            )

        if event_asset.status == "MISSING":
            append_activity(
                activities,
                "EVENT_ASSET_MISSING",
                "Asset mancante in evento",
                f"{inventory_code} - {event_name}",
                event_asset.created_at,
                "event",
                "critical",
                {
                    "event_asset_id": event_asset.id,
                    "event_id": event_asset.event_id,
                    "asset_id": event_asset.asset_id,
                    "inventory_code": asset.inventory_code if asset else None,
                    "href": f"/events?eventId={event_asset.event_id}",
                },
            )

    event_stocks = limited(
        db.query(EventStock).order_by(EventStock.created_at.desc())
    ).all()

    for event_stock in event_stocks:
        event = events.get(event_stock.event_id)
        stock = stocks.get(event_stock.stock_card_id)
        item = items.get(stock.item_id) if stock else None
        event_name = event.name if event else f"Evento ID {event_stock.event_id}"
        item_label = item_export_label(item) or f"Stockcard ID {event_stock.stock_card_id}"

        append_activity(
            activities,
            "EVENT_STOCK_OUT",
            "Stock uscito per evento",
            f"{item_label} -> {event_name} ({event_stock.quantity_out} pezzi)",
            event_stock.created_at,
            "event",
            "success",
            {
                "event_stock_id": event_stock.id,
                "event_id": event_stock.event_id,
                "stock_card_id": event_stock.stock_card_id,
                "href": f"/events?eventId={event_stock.event_id}",
            },
        )

    recent_events = limited(
        db.query(Event).order_by(Event.created_at.desc())
    ).all()

    for event in recent_events:
        append_activity(
            activities,
            "EVENT_CREATE",
            "Evento creato",
            event.name,
            event.created_at,
            "event",
            "success",
            {
                "event_id": event.id,
                "event_status": event.status,
                "href": f"/events?eventId={event.id}",
            },
        )

    return sorted(
        activities,
        key=lambda activity: activity["timestamp"],
        reverse=True,
    )


def filter_activity_log(
    activities: list[dict],
    q: str | None = None,
    category: str | None = None,
    severity: str | None = None,
):
    normalized_query = q.strip().lower() if q else None
    normalized_category = normalize_activity_filter(category)
    normalized_severity = severity.strip().lower() if severity else None

    result = []

    for activity in activities:
        if normalized_category and activity["category"] != normalized_category:
            continue

        if normalized_severity and activity["severity"] != normalized_severity:
            continue

        if normalized_query and normalized_query not in activity_search_blob(activity):
            continue

        result.append(activity)

    return result


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/locations")
def get_locations():
    db = SessionLocal()
    rows = db.query(Location).all()
    db.close()
    return rows


@app.get("/categories")
def get_categories():
    db = SessionLocal()
    rows = db.query(Category).order_by(Category.name.asc()).all()
    db.close()
    return rows


@app.get("/alerts")
def get_alerts():
    db = SessionLocal()

    try:
        critical = []
        warning = []

        missing_assets = (
            db.query(Asset)
            .filter(Asset.status == "MANCANTE")
            .order_by(Asset.inventory_code.asc())
            .all()
        )

        for asset in missing_assets:
            critical.append({
                "type": "ASSET_MISSING",
                "message": f"Asset mancante: {asset.inventory_code}",
                "references": {
                    "asset_id": asset.id,
                    "inventory_code": asset.inventory_code,
                    "href": f"/assets/{asset.inventory_code}",
                },
            })

        open_events = (
            db.query(Event)
            .filter(Event.status == "OPEN")
            .order_by(Event.created_at.asc())
            .all()
        )

        for event in open_events:
            open_assets = (
                db.query(EventAsset)
                .filter(
                    EventAsset.event_id == event.id,
                    EventAsset.status == "OUT",
                )
                .count()
            )

            if open_assets > 0:
                critical.append({
                    "type": "EVENT_ASSETS_OUT",
                    "message": f"Evento aperto con {open_assets} asset ancora fuori: {event.name}",
                    "references": {
                        "event_id": event.id,
                        "event_name": event.name,
                        "assets_out": open_assets,
                        "href": f"/events?eventId={event.id}",
                    },
                })

        low_stocks = (
            db.query(StockCard)
            .filter(StockCard.quantity <= StockCard.min_threshold)
            .order_by(StockCard.quantity.asc())
            .all()
        )

        for stock in low_stocks:
            item = db.execute(select(Item).where(Item.id == stock.item_id)).scalar_one_or_none()
            location = db.execute(select(Location).where(Location.id == stock.location_id)).scalar_one_or_none()

            item_label = item.name if item else f"Item ID {stock.item_id}"
            location_label = location.code if location else f"Sede ID {stock.location_id}"

            warning.append({
                "type": "LOW_STOCK",
                "message": (
                    f"Stock sotto soglia: {item_label} ({location_label}) "
                    f"- disponibili {stock.quantity}, soglia {stock.min_threshold}"
                ),
                "references": {
                    "stock_card_id": stock.id,
                    "item_id": stock.item_id,
                    "location_id": stock.location_id,
                    "quantity": stock.quantity,
                    "min_threshold": stock.min_threshold,
                    "href": "/stocks?lowStock=1",
                },
            })

        stale_threshold = datetime.utcnow() - timedelta(days=7)
        stale_events = (
            db.query(Event)
            .filter(
                Event.status == "OPEN",
                Event.created_at < stale_threshold,
            )
            .order_by(Event.created_at.asc())
            .all()
        )

        for event in stale_events:
            age_days = (datetime.utcnow() - event.created_at).days

            warning.append({
                "type": "EVENT_OPEN_TOO_LONG",
                "message": f"Evento aperto da {age_days} giorni: {event.name}",
                "references": {
                    "event_id": event.id,
                    "event_name": event.name,
                    "age_days": age_days,
                    "href": f"/events?eventId={event.id}",
                },
            })

        return {
            "critical": critical,
            "warning": warning,
        }
    finally:
        db.close()


@app.get("/dashboard/activity")
def get_dashboard_activity():
    db = SessionLocal()

    try:
        return build_activity_log(db, source_limit=40)[:20]
    finally:
        db.close()


@app.get("/activity")
def get_activity(
    q: str | None = None,
    category: str | None = None,
    severity: str | None = None,
    limit: int = 50,
    offset: int = 0,
):
    db = SessionLocal()

    try:
        safe_limit = min(max(limit, 1), 100)
        safe_offset = max(offset, 0)

        activities = build_activity_log(db)
        filtered_activities = filter_activity_log(
            activities,
            q=q,
            category=category,
            severity=severity,
        )

        total = len(filtered_activities)
        items = filtered_activities[safe_offset:safe_offset + safe_limit]

        return {
            "items": items,
            "total": total,
            "limit": safe_limit,
            "offset": safe_offset,
            "has_more": safe_offset + safe_limit < total,
        }
    finally:
        db.close()


@app.get("/dashboard/locations")
def get_dashboard_locations():
    db = SessionLocal()

    try:
        locations = db.query(Location).order_by(Location.name.asc()).all()
        result = []

        for location in locations:
            coordinates = LOCATION_COORDINATES.get(location.code)
            assets = db.query(Asset).filter(Asset.current_location_id == location.id).all()
            stockcards = db.query(StockCard).filter(StockCard.location_id == location.id).all()

            asset_total = len(assets)
            asset_in_location = sum(1 for asset in assets if asset.status == "IN_SEDE")
            asset_assigned = sum(1 for asset in assets if asset.status == "ASSEGNATO")
            asset_in_event = sum(1 for asset in assets if asset.status == "IN_EVENTO")
            asset_missing = sum(1 for asset in assets if asset.status == "MANCANTE")
            stock_quantity_total = sum(stock.quantity for stock in stockcards)
            low_stock_count = sum(
                1 for stock in stockcards if stock.quantity <= stock.min_threshold
            )

            location_key = normalize_import_key(location.code)
            location_name_key = normalize_import_key(location.name)
            open_events_count = 0
            for event in db.query(Event).filter(Event.status == "OPEN").all():
                event_location = normalize_import_key(event.location)
                if event_location and (
                    location_key in event_location
                    or location_name_key in event_location
                    or event_location in location_name_key
                ):
                    open_events_count += 1

            alert_count = asset_missing + low_stock_count + open_events_count
            alert_level = "none"
            if asset_missing > 0:
                alert_level = "critical"
            elif low_stock_count > 0 or open_events_count > 0:
                alert_level = "warning"

            result.append({
                "location_id": location.id,
                "code": location.code,
                "name": location.name,
                "lat": coordinates["lat"] if coordinates else None,
                "lng": coordinates["lng"] if coordinates else None,
                "asset_total": asset_total,
                "asset_in_location": asset_in_location,
                "asset_assigned": asset_assigned,
                "asset_in_event": asset_in_event,
                "asset_missing": asset_missing,
                "stockcard_count": len(stockcards),
                "stock_quantity_total": stock_quantity_total,
                "low_stock_count": low_stock_count,
                "open_events_count": open_events_count,
                "alert_count": alert_count,
                "alert_level": alert_level,
            })

        return result
    finally:
        db.close()


@app.get("/search")
def global_search(q: str = ""):
    query = q.strip()
    normalized_query = query.lower()

    if len(normalized_query) < 2:
        return empty_search_response(query)

    db = SessionLocal()

    try:
        categories = {category.id: category for category in db.query(Category).all()}
        locations = {location.id: location for location in db.query(Location).all()}
        items = {item.id: item for item in db.query(Item).all()}

        results = {
            "assets": [],
            "items": [],
            "stocks": [],
            "events": [],
        }

        for asset in db.query(Asset).order_by(Asset.id.desc()).all():
            if len(results["assets"]) >= 5:
                break

            item = items.get(asset.item_id)
            category = categories.get(item.category_id) if item else None
            location = locations.get(asset.current_location_id)
            assignee = None
            if asset.assignee_id:
                assignee = db.execute(
                    select(Assignee).where(Assignee.id == asset.assignee_id)
                ).scalar_one_or_none()

            if not text_matches(
                normalized_query,
                asset.inventory_code,
                asset.notes,
                asset.status,
                asset.assigned_to,
                assignee.name if assignee else None,
                assignee.email if assignee else None,
                item.name if item else None,
                item.brand if item else None,
                item.model if item else None,
                category.name if category else None,
            ):
                continue

            item_label = item_export_label(item) or "Item non trovato"
            location_label = location_export_label(location) or "Sede non trovata"

            results["assets"].append({
                "type": "asset",
                "title": asset.inventory_code,
                "description": f"{item_label} - {asset.status} - {location_label}",
                "href": f"/assets/{asset.inventory_code}",
                "metadata": {
                    "asset_id": asset.id,
                    "inventory_code": asset.inventory_code,
                    "status": asset.status,
                    "assigned_to": asset.assigned_to,
                    "item_id": asset.item_id,
                    "location_id": asset.current_location_id,
                },
            })

        for item in db.query(Item).order_by(Item.id.desc()).all():
            if len(results["items"]) >= 5:
                break

            category = categories.get(item.category_id)

            if not text_matches(
                normalized_query,
                item.name,
                category.name if category else None,
                item.brand,
                item.model,
            ):
                continue

            details = [
                value
                for value in [
                    category.name if category else None,
                    item.brand,
                    item.model,
                    "serializzato" if item.is_serialized else "consumabile",
                ]
                if value
            ]

            results["items"].append({
                "type": "item",
                "title": item.name,
                "description": " - ".join(details),
                "href": "/items",
                "metadata": {
                    "item_id": item.id,
                    "category_id": item.category_id,
                    "is_serialized": item.is_serialized,
                },
            })

        for stock in db.query(StockCard).order_by(StockCard.id.desc()).all():
            if len(results["stocks"]) >= 5:
                break

            item = items.get(stock.item_id)
            category = categories.get(item.category_id) if item else None
            location = locations.get(stock.location_id)

            if not text_matches(
                normalized_query,
                item.name if item else None,
                item.brand if item else None,
                item.model if item else None,
                category.name if category else None,
                location.code if location else None,
                location.name if location else None,
                stock.notes,
            ):
                continue

            item_label = item_export_label(item) or f"Stockcard {stock.id}"
            location_label = location_export_label(location) or "Sede non trovata"

            results["stocks"].append({
                "type": "stock",
                "title": item_label,
                "description": f"{location_label} - disponibili {stock.quantity}, soglia {stock.min_threshold}",
                "href": "/stocks",
                "metadata": {
                    "stock_card_id": stock.id,
                    "item_id": stock.item_id,
                    "location_id": stock.location_id,
                    "quantity": stock.quantity,
                    "min_threshold": stock.min_threshold,
                },
            })

        for event in db.query(Event).order_by(Event.created_at.desc()).all():
            if len(results["events"]) >= 5:
                break

            if not text_matches(
                normalized_query,
                event.name,
                event.notes,
                event.location,
                event.status,
                event.manager,
            ):
                continue

            details = [
                value
                for value in [
                    event.location,
                    event.status,
                    event.start_date,
                    event.end_date,
                ]
                if value
            ]

            results["events"].append({
                "type": "event",
                "title": event.name,
                "description": " - ".join(details),
                "href": f"/events?eventId={event.id}",
                "metadata": {
                    "event_id": event.id,
                    "status": event.status,
                    "location": event.location,
                },
            })

        return {
            "query": query,
            "results": results,
        }
    finally:
        db.close()


@app.get("/imports/template.xlsx")
def import_template_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = IMPORT_SHEET_NAME
    sheet.append(IMPORT_COLUMNS)
    sheet.append([
        "ASSET",
        "LE1",
        "Informatica",
        "Notebook didattico",
        "Lenovo",
        "ThinkPad E14",
        "SI",
        2,
        "",
        "Asset iniziale da import",
        "Mario Rossi",
    ])
    sheet.append([
        "STOCK",
        "LE1",
        "Consumabili",
        "Quaderno A4",
        "",
        "",
        "NO",
        50,
        10,
        "Dotazione iniziale",
        "",
    ])

    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 14,
        "B": 12,
        "C": 24,
        "D": 28,
        "E": 18,
        "F": 24,
        "G": 14,
        "H": 12,
        "I": 16,
        "J": 36,
        "K": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    instructions = workbook.create_sheet("Istruzioni")
    instructions.append(["Campo", "Descrizione"])
    instructions.append(["tipo", "ASSET per beni serializzati, STOCK per consumabili"])
    instructions.append(["sede", "Codice sede esistente, es. LE1"])
    instructions.append(["categoria", "Categoria da riusare o creare"])
    instructions.append(["nome", "Nome item da riusare o creare"])
    instructions.append(["serializzato", "SI per ASSET, NO per STOCK"])
    instructions.append(["quantita", "Numero asset da creare o pezzi stock da caricare"])
    instructions.append(["soglia_minima", "Solo per STOCK, opzionale"])
    instructions.append(["assegnatario", "Solo per ASSET, crea/riusa un assegnatario PERSON"])
    for cell in instructions[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 72

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template-import-inventario.xlsx"},
    )


async def read_import_upload(file: UploadFile):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Carica un file .xlsx valido.")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="File vuoto o non leggibile.")

    return file_bytes


@app.post("/imports/preview")
async def import_preview(file: UploadFile = File(...)):
    file_bytes = await read_import_upload(file)
    db = SessionLocal()

    try:
        return build_import_preview(file_bytes, db)
    finally:
        db.close()


@app.post("/imports/commit")
async def import_commit(file: UploadFile = File(...)):
    file_bytes = await read_import_upload(file)
    db = SessionLocal()

    try:
        plan = build_import_preview(file_bytes, db)
        if not plan["can_commit"]:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Import non confermabile: correggi gli errori in preview.",
                    "preview": plan,
                },
            )

        committed = commit_import_operations(plan, db)
        db.commit()

        return {
            "committed": True,
            "summary": plan["summary"],
            "result": committed,
        }
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.get("/exports/assets.xlsx")
def export_assets_xlsx():
    from openpyxl import Workbook

    db = SessionLocal()

    try:
        workbook = Workbook()
        assets_sheet = workbook.active
        assets_sheet.title = "Asset"
        assets_sheet.append([
            "ID",
            "Codice inventario",
            "Item",
            "Categoria",
            "Marca",
            "Modello",
            "Sede attuale",
            "Stato",
            "Assegnato a",
            "Note",
        ])

        items = {item.id: item for item in db.query(Item).all()}
        locations = {location.id: location for location in db.query(Location).all()}
        assets = db.query(Asset).order_by(Asset.inventory_code.asc()).all()

        for asset in assets:
            item = items.get(asset.item_id)
            category = item.category.name if item and item.category else ""

            assets_sheet.append([
                asset.id,
                asset.inventory_code,
                item_export_label(item),
                category,
                item.brand if item else "",
                item.model if item else "",
                location_export_label(locations.get(asset.current_location_id)),
                asset.status,
                asset.assigned_to or "",
                asset.notes or "",
            ])

        movements_sheet = workbook.create_sheet("Movimenti asset")
        movements_sheet.append([
            "ID",
            "Codice inventario",
            "Da sede",
            "A sede",
            "Data movimento",
            "Operatore",
            "Note",
        ])

        assets_by_id = {asset.id: asset for asset in assets}
        movements = (
            db.query(AssetMovement)
            .order_by(AssetMovement.moved_at.desc())
            .all()
        )

        for movement in movements:
            asset = assets_by_id.get(movement.asset_id)

            movements_sheet.append([
                movement.id,
                asset.inventory_code if asset else "",
                location_export_label(locations.get(movement.from_location_id)),
                location_export_label(locations.get(movement.to_location_id)),
                movement.moved_at,
                movement.moved_by or "",
                movement.notes or "",
            ])

        logs_sheet = workbook.create_sheet("Log asset")
        logs_sheet.append([
            "ID",
            "Codice inventario",
            "Tipo azione",
            "Descrizione",
            "Data",
            "Operatore",
        ])

        logs = (
            db.query(AssetLog)
            .order_by(AssetLog.created_at.desc())
            .all()
        )

        for log in logs:
            asset = assets_by_id.get(log.asset_id)

            logs_sheet.append([
                log.id,
                asset.inventory_code if asset else "",
                log.action_type,
                log.description,
                log.created_at,
                log.created_by or "",
            ])

        for sheet in workbook.worksheets:
            style_export_sheet(sheet)

        return export_workbook(workbook, "inventario-assets.xlsx")
    finally:
        db.close()


@app.get("/exports/stocks.xlsx")
def export_stocks_xlsx():
    from openpyxl import Workbook

    db = SessionLocal()

    try:
        workbook = Workbook()
        stocks_sheet = workbook.active
        stocks_sheet.title = "Stockcard"
        stocks_sheet.append([
            "ID",
            "Item",
            "Categoria",
            "Marca",
            "Modello",
            "Sede",
            "Quantita disponibile",
            "Soglia minima",
            "Sotto soglia",
            "Note",
        ])

        items = {item.id: item for item in db.query(Item).all()}
        locations = {location.id: location for location in db.query(Location).all()}
        stocks = db.query(StockCard).order_by(StockCard.id.asc()).all()

        for stock in stocks:
            item = items.get(stock.item_id)
            category = item.category.name if item and item.category else ""

            stocks_sheet.append([
                stock.id,
                item_export_label(item),
                category,
                item.brand if item else "",
                item.model if item else "",
                location_export_label(locations.get(stock.location_id)),
                stock.quantity,
                stock.min_threshold,
                "SI" if stock.quantity <= stock.min_threshold else "NO",
                stock.notes or "",
            ])

        movements_sheet = workbook.create_sheet("Movimenti stock")
        movements_sheet.append([
            "ID",
            "Stockcard ID",
            "Item",
            "Sede",
            "Tipo movimento",
            "Quantita",
            "Data",
            "Note",
        ])

        stocks_by_id = {stock.id: stock for stock in stocks}
        movements = (
            db.query(StockMovement)
            .order_by(StockMovement.created_at.desc())
            .all()
        )

        for movement in movements:
            stock = stocks_by_id.get(movement.stock_card_id)
            item = items.get(stock.item_id) if stock else None
            location = locations.get(stock.location_id) if stock else None

            movements_sheet.append([
                movement.id,
                movement.stock_card_id,
                item_export_label(item),
                location_export_label(location),
                movement.movement_type,
                movement.quantity,
                movement.created_at,
                movement.notes or "",
            ])

        for sheet in workbook.worksheets:
            style_export_sheet(sheet)

        return export_workbook(workbook, "inventario-stock.xlsx")
    finally:
        db.close()


@app.get("/exports/events.xlsx")
def export_events_xlsx():
    from openpyxl import Workbook

    db = SessionLocal()

    try:
        workbook = Workbook()
        events_sheet = workbook.active
        events_sheet.title = "Eventi"
        events_sheet.append([
            "ID",
            "Nome",
            "Luogo",
            "Data inizio",
            "Data fine",
            "Referente",
            "Stato",
            "Creato il",
            "Asset collegati",
            "Stock collegati",
            "Note",
        ])

        events = db.query(Event).order_by(Event.created_at.desc()).all()

        for event in events:
            linked_assets = db.query(EventAsset).filter(EventAsset.event_id == event.id).count()
            linked_stocks = db.query(EventStock).filter(EventStock.event_id == event.id).count()

            events_sheet.append([
                event.id,
                event.name,
                event.location or "",
                event.start_date or "",
                event.end_date or "",
                event.manager or "",
                event.status,
                event.created_at,
                linked_assets,
                linked_stocks,
                event.notes or "",
            ])

        event_assets_sheet = workbook.create_sheet("Asset eventi")
        event_assets_sheet.append([
            "ID",
            "Evento",
            "Codice asset",
            "Stato",
            "Creato il",
            "Rientrato il",
            "Note",
        ])

        assets = {asset.id: asset for asset in db.query(Asset).all()}
        events_by_id = {event.id: event for event in events}
        event_assets = (
            db.query(EventAsset)
            .order_by(EventAsset.created_at.desc())
            .all()
        )

        for event_asset in event_assets:
            event = events_by_id.get(event_asset.event_id)
            asset = assets.get(event_asset.asset_id)

            event_assets_sheet.append([
                event_asset.id,
                event.name if event else "",
                asset.inventory_code if asset else "",
                event_asset.status,
                event_asset.created_at,
                event_asset.returned_at,
                event_asset.notes or "",
            ])

        event_stocks_sheet = workbook.create_sheet("Stock eventi")
        event_stocks_sheet.append([
            "ID",
            "Evento",
            "Stockcard ID",
            "Item",
            "Quantita uscita",
            "Quantita rientrata",
            "Da rientrare",
            "Creato il",
            "Note",
        ])

        stocks = {stock.id: stock for stock in db.query(StockCard).all()}
        items = {item.id: item for item in db.query(Item).all()}
        event_stocks = (
            db.query(EventStock)
            .order_by(EventStock.created_at.desc())
            .all()
        )

        for event_stock in event_stocks:
            event = events_by_id.get(event_stock.event_id)
            stock = stocks.get(event_stock.stock_card_id)
            item = items.get(stock.item_id) if stock else None

            event_stocks_sheet.append([
                event_stock.id,
                event.name if event else "",
                event_stock.stock_card_id,
                item_export_label(item),
                event_stock.quantity_out,
                event_stock.quantity_returned,
                max(0, event_stock.quantity_out - event_stock.quantity_returned),
                event_stock.created_at,
                event_stock.notes or "",
            ])

        for sheet in workbook.worksheets:
            style_export_sheet(sheet)

        return export_workbook(workbook, "inventario-eventi.xlsx")
    finally:
        db.close()


@app.post("/items")
def create_item(
    name: str = Body(...),
    category_id: int = Body(...),
    brand: str | None = Body(None),
    model: str | None = Body(None),
    technical_specs: str | None = Body(None),
    is_serialized: bool = Body(True),
):
    db = SessionLocal()
    category = db.execute(select(Category).where(Category.id == category_id)).scalar_one_or_none()
    if not category:
        db.close()
        raise HTTPException(status_code=404, detail=f"Categoria non trovata: {category_id}")
    item = Item(
        name=name,
        category_id=category.id,
        brand=brand,
        model=model,
        technical_specs=technical_specs,
        is_serialized=is_serialized,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    item.category = category
    serialized = serialize_item(item, asset_count=0, stock_card_count=0)
    db.close()
    return serialized



@app.get("/items")
def list_items():
    db = SessionLocal()
    items = db.query(Item).all()
    rows = []

    for item in items:
        asset_count = db.query(Asset).filter(Asset.item_id == item.id).count()
        stock_card_count = db.query(StockCard).filter(StockCard.item_id == item.id).count()
        rows.append(
            serialize_item(
                item,
                asset_count=asset_count,
                stock_card_count=stock_card_count,
            )
        )
    db.close()
    return rows


# --- GET/UPDATE ITEM ENDPOINTS ---

@app.get("/items/{item_id}")
def get_item(item_id: int):
    db = SessionLocal()
    item = db.execute(select(Item).where(Item.id == item_id)).scalar_one_or_none()

    if not item:
        db.close()
        raise HTTPException(status_code=404, detail=f"Item non trovato: {item_id}")

    asset_count = db.query(Asset).filter(Asset.item_id == item.id).count()
    stock_card_count = db.query(StockCard).filter(StockCard.item_id == item.id).count()
    serialized = serialize_item(
        item,
        asset_count=asset_count,
        stock_card_count=stock_card_count,
    )
    db.close()
    return serialized



@app.put("/items/{item_id}")
def update_item(
    item_id: int,
    name: str = Body(...),
    category_id: int = Body(...),
    brand: str | None = Body(None),
    model: str | None = Body(None),
    technical_specs: str | None = Body(None),
    is_serialized: bool = Body(True),
):
    db = SessionLocal()

    try:
        item = db.execute(select(Item).where(Item.id == item_id)).scalar_one_or_none()
        if not item:
            raise HTTPException(status_code=404, detail=f"Item non trovato: {item_id}")

        category = db.execute(select(Category).where(Category.id == category_id)).scalar_one_or_none()
        if not category:
            raise HTTPException(status_code=404, detail=f"Categoria non trovata: {category_id}")

        item.name = name
        item.category_id = category.id
        item.brand = brand
        item.model = model
        item.technical_specs = technical_specs
        item.is_serialized = is_serialized

        db.commit()
        db.refresh(item)
        item.category = category
        asset_count = db.query(Asset).filter(Asset.item_id == item.id).count()
        stock_card_count = db.query(StockCard).filter(StockCard.item_id == item.id).count()
        serialized = serialize_item(
            item,
            asset_count=asset_count,
            stock_card_count=stock_card_count,
        )
        return serialized
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# --- DELETE ITEM ENDPOINT ---

@app.delete("/items/{item_id}")
def delete_item(item_id: int):
    db = SessionLocal()

    try:
        item = db.execute(select(Item).where(Item.id == item_id)).scalar_one_or_none()
        if not item:
            raise HTTPException(status_code=404, detail=f"Item non trovato: {item_id}")

        asset_count = db.query(Asset).filter(Asset.item_id == item.id).count()
        if asset_count > 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Impossibile eliminare item: esistono asset collegati. "
                    f"Asset collegati: {asset_count}."
                ),
            )

        stock_count = db.query(StockCard).filter(StockCard.item_id == item.id).count()
        if stock_count > 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Impossibile eliminare item: esistono schede stock collegate. "
                    f"Stockcard collegate: {stock_count}."
                ),
            )

        db.delete(item)
        db.commit()

        return {"deleted": True, "item_id": item_id}
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.get("/assignees")
def list_assignees():
    db = SessionLocal()

    try:
        assignees = db.query(Assignee).order_by(Assignee.name.asc()).all()
        return [
            serialize_assignee(
                assignee,
                asset_count=db.query(Asset).filter(Asset.assignee_id == assignee.id).count(),
            )
            for assignee in assignees
        ]
    finally:
        db.close()


@app.get("/assignees/{assignee_id}")
def get_assignee(assignee_id: int):
    db = SessionLocal()

    try:
        assignee = db.execute(
            select(Assignee).where(Assignee.id == assignee_id)
        ).scalar_one_or_none()

        if not assignee:
            raise HTTPException(status_code=404, detail="Assegnatario non trovato")

        assets = (
            db.query(Asset)
            .filter(Asset.assignee_id == assignee.id)
            .order_by(Asset.inventory_code.asc())
            .all()
        )

        return serialize_assignee(
            assignee,
            asset_count=len(assets),
            assets=assets,
        )
    finally:
        db.close()


@app.post("/assignees")
def create_assignee(
    name: str = Body(...),
    type: str = Body(...),
    email: str | None = Body(None),
    phone: str | None = Body(None),
    notes: str | None = Body(None),
    is_active: bool = Body(True),
):
    db = SessionLocal()

    try:
        assignee = Assignee(
            name=name.strip(),
            type=validate_assignee_type(type),
            email=email.strip() if email else None,
            phone=phone.strip() if phone else None,
            notes=notes.strip() if notes else None,
            is_active=is_active,
        )

        if not assignee.name:
            raise HTTPException(status_code=400, detail="Nome assegnatario obbligatorio.")

        db.add(assignee)
        db.commit()
        db.refresh(assignee)

        return serialize_assignee(assignee, asset_count=0)
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.put("/assignees/{assignee_id}")
def update_assignee(
    assignee_id: int,
    name: str = Body(...),
    type: str = Body(...),
    email: str | None = Body(None),
    phone: str | None = Body(None),
    notes: str | None = Body(None),
    is_active: bool = Body(True),
):
    db = SessionLocal()

    try:
        assignee = db.execute(
            select(Assignee).where(Assignee.id == assignee_id)
        ).scalar_one_or_none()

        if not assignee:
            raise HTTPException(status_code=404, detail="Assegnatario non trovato")

        clean_name = name.strip()
        if not clean_name:
            raise HTTPException(status_code=400, detail="Nome assegnatario obbligatorio.")

        assignee.name = clean_name
        assignee.type = validate_assignee_type(type)
        assignee.email = email.strip() if email else None
        assignee.phone = phone.strip() if phone else None
        assignee.notes = notes.strip() if notes else None
        assignee.is_active = is_active

        linked_assets = db.query(Asset).filter(Asset.assignee_id == assignee.id).all()
        for asset in linked_assets:
            asset.assigned_to = assignee.name

        db.commit()
        db.refresh(assignee)

        return serialize_assignee(
            assignee,
            asset_count=len(linked_assets),
            assets=linked_assets,
        )
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.delete("/assignees/{assignee_id}")
def delete_assignee(assignee_id: int):
    db = SessionLocal()

    try:
        assignee = db.execute(
            select(Assignee).where(Assignee.id == assignee_id)
        ).scalar_one_or_none()

        if not assignee:
            raise HTTPException(status_code=404, detail="Assegnatario non trovato")

        linked_assets = db.query(Asset).filter(Asset.assignee_id == assignee.id).all()
        if linked_assets:
            assignee.is_active = False
            db.commit()
            db.refresh(assignee)

            return {
                "deleted": False,
                "deactivated": True,
                "assignee_id": assignee.id,
                "detail": "Assegnatario disattivato perché collegato ad asset.",
            }

        db.delete(assignee)
        db.commit()

        return {
            "deleted": True,
            "deactivated": False,
            "assignee_id": assignee_id,
        }
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.post("/assets")
def create_asset(item_id: int, location_code: str, notes: str | None = None):
    db = SessionLocal()

    item = db.execute(select(Item).where(Item.id == item_id)).scalar_one_or_none()
    if not item:
        db.close()
        raise HTTPException(status_code=404, detail=f"Item non trovato: {item_id}")
    if not item.is_serialized:
        db.close()
        raise HTTPException(status_code=400, detail="Questo item non è serializzato (usa StockCard in futuro).")

    # Transazione semplice: se qualcosa va male, rollback
    try:
        inv_code = generate_inventory_code(db, location_code)
        loc = db.execute(select(Location).where(Location.code == location_code)).scalar_one()

        asset = Asset(
            inventory_code=inv_code,
            item_id=item.id,
            current_location_id=loc.id,
            status="IN_SEDE",
            notes=notes
        )
        db.add(asset)
        db.flush()
        create_asset_log(
            db,
            asset.id,
            "CREATE",
            f"Asset creato in sede {loc.code} - {loc.name}",
        )
        db.commit()
        db.refresh(asset)
        return asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.post("/assets/{asset_id}/transfer")
def transfer_asset(
    asset_id: int,
    to_location_code: str = Body(...),
    moved_by: str | None = Body(None),
    notes: str | None = Body(None),
):
    db = SessionLocal()
    try:
        asset = db.execute(select(Asset).where(Asset.id == asset_id)).scalar_one_or_none()
        if not asset:
            raise HTTPException(status_code=404, detail=f"Asset non trovato: {asset_id}")

        to_loc = db.execute(select(Location).where(Location.code == to_location_code)).scalar_one_or_none()
        if not to_loc:
            raise HTTPException(status_code=404, detail=f"Sede non trovata: {to_location_code}")

        from_loc_id = asset.current_location_id

        # Se vuoi evitare trasferimenti "uguali"
        if from_loc_id == to_loc.id:
            raise HTTPException(status_code=400, detail="L'asset è già in questa sede.")

        movement = AssetMovement(
            asset_id=asset.id,
            from_location_id=from_loc_id,
            to_location_id=to_loc.id,
            moved_by=moved_by,
            notes=notes
        )
        db.add(movement)

        # aggiorna la sede attuale dell'asset
        asset.current_location_id = to_loc.id
        asset.status = "IN_SEDE"

        create_asset_log(
            db,
            asset.id,
            "TRANSFER",
            f"Asset trasferito alla sede {to_loc.code} - {to_loc.name}",
            created_by=moved_by,
        )

        db.commit()
        db.refresh(asset)
        return asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

@app.post("/assets/{asset_id}/assign")
def assign_asset(
    asset_id: int,
    assigned_to: str | None = Body(None),
    assignee_id: int | None = Body(None),
    notes: str | None = Body(None)
):
    db = SessionLocal()

    try:
        asset = db.query(Asset).filter(Asset.id == asset_id).first()

        if not asset:
            raise HTTPException(status_code=404, detail="Asset non trovato")

        assignee = None
        if assignee_id is not None:
            assignee = db.execute(
                select(Assignee).where(Assignee.id == assignee_id)
            ).scalar_one_or_none()

            if not assignee:
                raise HTTPException(status_code=404, detail="Assegnatario non trovato")

            if not assignee.is_active:
                raise HTTPException(status_code=400, detail="Assegnatario non attivo")

            asset.assignee_id = assignee.id
            asset.assigned_to = assignee.name
        elif assigned_to and assigned_to.strip():
            asset.assignee_id = None
            asset.assigned_to = assigned_to.strip()
        else:
            raise HTTPException(
                status_code=400,
                detail="Indica un assegnatario strutturato o un testo assegnazione.",
            )

        asset.status = "ASSEGNATO"

        if notes:
            asset.notes = notes

        create_asset_log(
            db,
            asset.id,
            "ASSIGN",
            f"Asset assegnato a {asset.assigned_to}",
        )

        db.commit()
        db.refresh(asset)

        return asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()



# Unassign endpoint
@app.post("/assets/{asset_id}/unassign")
def unassign_asset(asset_id: int):
    db = SessionLocal()

    try:
        asset = db.query(Asset).filter(Asset.id == asset_id).first()

        if not asset:
            raise HTTPException(status_code=404, detail="Asset non trovato")

        previous_assignee = asset.assigned_to
        asset.assignee_id = None
        asset.assigned_to = None
        asset.status = "IN_SEDE"

        create_asset_log(
            db,
            asset.id,
            "UNASSIGN",
            "Assegnazione rimossa" if not previous_assignee else f"Assegnazione rimossa da {previous_assignee}",
        )

        db.commit()
        db.refresh(asset)

        return asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# --- MARK ASSET MISSING ENDPOINT ---


@app.post("/assets/{asset_id}/missing")
def mark_asset_missing(
    asset_id: int,
    notes: str | None = Body(None),
):
    db = SessionLocal()

    try:
        asset = db.query(Asset).filter(Asset.id == asset_id).first()

        if not asset:
            raise HTTPException(status_code=404, detail="Asset non trovato")

        asset.status = "MANCANTE"

        if notes:
            asset.notes = notes

        create_asset_log(
            db,
            asset.id,
            "MARK_MISSING",
            "Asset segnato come mancante",
        )

        db.commit()
        db.refresh(asset)

        return asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# --- RESTORE ASSET ENDPOINT ---

@app.post("/assets/{asset_id}/restore")
def restore_asset(asset_id: int):
    db = SessionLocal()

    try:
        asset = db.query(Asset).filter(Asset.id == asset_id).first()

        if not asset:
            raise HTTPException(status_code=404, detail="Asset non trovato")

        if asset.status != "MANCANTE":
            raise HTTPException(
                status_code=400,
                detail="Solo un asset mancante può essere ripristinato.",
            )

        asset.status = "IN_SEDE"

        create_asset_log(
            db,
            asset.id,
            "RESTORE",
            "Asset ripristinato e riportato in sede",
        )

        db.commit()
        db.refresh(asset)

        return asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

@app.get("/assets")
def list_assets():
    db = SessionLocal()
    rows = db.query(Asset).all()
    db.close()
    return rows

@app.get("/assets/{inventory_code}")
def get_asset_by_code(inventory_code: str):
    db = SessionLocal()
    asset = db.execute(
        select(Asset).where(Asset.inventory_code == inventory_code)
    ).scalar_one_or_none()
    db.close()

    if not asset:
        raise HTTPException(status_code=404, detail="Asset non trovato")
    return asset


@app.get("/assets/{inventory_code}/detail")
def get_asset_detail_by_code(inventory_code: str):
    db = SessionLocal()
    asset = db.execute(
        select(Asset).where(Asset.inventory_code == inventory_code)
    ).scalar_one_or_none()

    if not asset:
        db.close()
        raise HTTPException(status_code=404, detail="Asset non trovato")

    item = db.execute(select(Item).where(Item.id == asset.item_id)).scalar_one_or_none()
    location = db.execute(
        select(Location).where(Location.id == asset.current_location_id)
    ).scalar_one_or_none()
    assignee = None
    if asset.assignee_id:
        assignee = db.execute(
            select(Assignee).where(Assignee.id == asset.assignee_id)
        ).scalar_one_or_none()

    serialized_item = None if not item else serialize_item(
        item,
        asset_count=db.query(Asset).filter(Asset.item_id == item.id).count(),
        stock_card_count=db.query(StockCard).filter(StockCard.item_id == item.id).count(),
    )

    result = {
        "asset": asset,
        "item": serialized_item,
        "location": location,
        "assignee": None if not assignee else serialize_assignee(
            assignee,
            asset_count=db.query(Asset).filter(Asset.assignee_id == assignee.id).count(),
        ),
    }

    db.close()
    return result

@app.get("/assets/{inventory_code}/qr")
def get_asset_qr(inventory_code: str):
    db = SessionLocal()

    asset = db.query(Asset).filter(Asset.inventory_code == inventory_code).first()
    db.close()

    if not asset:
        raise HTTPException(status_code=404, detail="Asset non trovato")

    url = f"http://localhost:8000/assets/{inventory_code}"

    qr = qrcode.make(url)

    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    buffer.seek(0)

    return StreamingResponse(buffer, media_type="image/png")

@app.get("/assets/{asset_id}/history")
def asset_history(asset_id: int):
    db = SessionLocal()
    rows = (
        db.query(AssetMovement)
        .filter(AssetMovement.asset_id == asset_id)
        .order_by(AssetMovement.moved_at.desc())
        .all()
    )
    db.close()
    return rows


@app.get("/assets/{asset_id}/logs")
def asset_logs(asset_id: int):
    db = SessionLocal()
    rows = (
        db.query(AssetLog)
        .filter(AssetLog.asset_id == asset_id)
        .order_by(AssetLog.created_at.desc())
        .all()
    )
    db.close()
    return rows

@app.get("/debug/routes")
def debug_routes():
    return [r.path for r in app.routes]

@app.get("/ping")
def ping():
    return {"pong": True}


@app.get("/assets-search")
def search_assets(q: str):
    q = q.strip()
    db = SessionLocal()
    rows = db.execute(
        select(Asset).where(
            or_(
                Asset.inventory_code.ilike(f"%{q}%"),
                Asset.assigned_to.ilike(f"%{q}%"),
                Asset.notes.ilike(f"%{q}%"),
            )
        )
    ).scalars().all()
    db.close()
    return rows


# --- STOCKS ENDPOINTS ---

@app.get("/stocks")
def list_stocks():
    db = SessionLocal()
    rows = db.query(StockCard).all()
    db.close()
    return rows


@app.post("/stocks")
def create_stock_card(
    item_id: int = Body(...),
    location_code: str = Body(...),
    quantity: int = Body(0),
    min_threshold: int = Body(0),
    notes: str | None = Body(None),
):
    db = SessionLocal()

    try:
        item = db.execute(select(Item).where(Item.id == item_id)).scalar_one_or_none()
        if not item:
            raise HTTPException(status_code=404, detail=f"Item non trovato: {item_id}")

        if item.is_serialized:
            raise HTTPException(
                status_code=400,
                detail="Questo item è serializzato: usa /assets per creare beni inventariati singolarmente."
            )

        location = db.execute(select(Location).where(Location.code == location_code)).scalar_one_or_none()
        if not location:
            raise HTTPException(status_code=404, detail=f"Sede non trovata: {location_code}")

        existing = db.execute(
            select(StockCard).where(
                StockCard.item_id == item.id,
                StockCard.location_id == location.id,
            )
        ).scalar_one_or_none()

        if existing:
            raise HTTPException(
                status_code=400,
                detail="Esiste già una scheda stock per questo item in questa sede."
            )

        if quantity < 0:
            raise HTTPException(status_code=400, detail="La quantità iniziale non può essere negativa.")

        if min_threshold < 0:
            raise HTTPException(status_code=400, detail="La soglia minima non può essere negativa.")

        stock = StockCard(
            item_id=item.id,
            location_id=location.id,
            quantity=quantity,
            min_threshold=min_threshold,
            notes=notes,
        )
        db.add(stock)
        db.flush()

        if quantity > 0:
            movement = StockMovement(
                stock_card_id=stock.id,
                movement_type="LOAD",
                quantity=quantity,
                notes="Carico iniziale",
            )
            db.add(movement)

        db.commit()
        db.refresh(stock)
        return stock
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.post("/stocks/{stock_id}/movement")
def create_stock_movement(
    stock_id: int,
    movement_type: str = Body(...),
    quantity: int = Body(...),
    notes: str | None = Body(None),
):
    db = SessionLocal()

    try:
        stock = db.execute(select(StockCard).where(StockCard.id == stock_id)).scalar_one_or_none()
        if not stock:
            raise HTTPException(status_code=404, detail=f"Scheda stock non trovata: {stock_id}")

        movement_type = movement_type.upper().strip()
        allowed = {"LOAD", "UNLOAD", "RETURN", "ADJUST"}
        if movement_type not in allowed:
            raise HTTPException(
                status_code=400,
                detail="Tipo movimento non valido. Usa LOAD, UNLOAD, RETURN o ADJUST."
            )

        if quantity < 0:
            raise HTTPException(status_code=400, detail="La quantità non può essere negativa.")

        if movement_type in {"LOAD", "RETURN"}:
            stock.quantity += quantity
        elif movement_type == "UNLOAD":
            if quantity > stock.quantity:
                raise HTTPException(
                    status_code=400,
                    detail="Quantità insufficiente per lo scarico."
                )
            stock.quantity -= quantity
        elif movement_type == "ADJUST":
            stock.quantity = quantity

        movement = StockMovement(
            stock_card_id=stock.id,
            movement_type=movement_type,
            quantity=quantity,
            notes=notes,
        )
        db.add(movement)

        db.commit()
        db.refresh(stock)
        return stock
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.get("/stocks/{stock_id}/history")
def stock_history(stock_id: int):
    db = SessionLocal()
    rows = (
        db.query(StockMovement)
        .filter(StockMovement.stock_card_id == stock_id)
        .order_by(StockMovement.created_at.desc())
        .all()
    )
    db.close()
    return rows


# --- EVENTS ENDPOINTS ---

@app.get("/events")
def list_events():
    db = SessionLocal()
    rows = db.query(Event).order_by(Event.created_at.desc()).all()
    db.close()
    return rows


@app.post("/events")
def create_event(
    name: str = Body(...),
    location: str | None = Body(None),
    start_date: str | None = Body(None),
    end_date: str | None = Body(None),
    manager: str | None = Body(None),
    notes: str | None = Body(None),
):
    db = SessionLocal()

    try:
        event = Event(
            name=name,
            location=location,
            start_date=start_date,
            end_date=end_date,
            manager=manager,
            notes=notes,
            status="OPEN",
        )
        db.add(event)
        db.commit()
        db.refresh(event)
        return event
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.get("/events/{event_id}")
def get_event(event_id: int):
    db = SessionLocal()

    event = db.execute(select(Event).where(Event.id == event_id)).scalar_one_or_none()
    if not event:
        db.close()
        raise HTTPException(status_code=404, detail="Evento non trovato")

    assets = db.query(EventAsset).filter(EventAsset.event_id == event_id).all()
    stocks = db.query(EventStock).filter(EventStock.event_id == event_id).all()

    db.close()

    return {
        "event": event,
        "assets": assets,
        "stocks": stocks,
    }


@app.post("/events/{event_id}/assets")
def add_asset_to_event(
    event_id: int,
    asset_id: int = Body(...),
    notes: str | None = Body(None),
):
    db = SessionLocal()

    try:
        event = db.execute(select(Event).where(Event.id == event_id)).scalar_one_or_none()
        if not event:
            raise HTTPException(status_code=404, detail="Evento non trovato")

        asset = db.execute(select(Asset).where(Asset.id == asset_id)).scalar_one_or_none()
        if not asset:
            raise HTTPException(status_code=404, detail="Asset non trovato")

        existing = db.execute(
            select(EventAsset).where(
                EventAsset.event_id == event_id,
                EventAsset.asset_id == asset_id,
            )
        ).scalar_one_or_none()

        if existing:
            raise HTTPException(status_code=400, detail="Asset già collegato a questo evento")

        event_asset = EventAsset(
            event_id=event.id,
            asset_id=asset.id,
            status="OUT",
            notes=notes,
        )
        db.add(event_asset)

        asset.status = "IN_EVENTO"

        create_asset_log(
            db,
            asset.id,
            "EVENT_OUT",
            f"Asset collegato all'evento: {event.name}",
        )

        db.commit()
        db.refresh(event_asset)
        return event_asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.post("/events/{event_id}/stocks")
def add_stock_to_event(
    event_id: int,
    stock_card_id: int = Body(...),
    quantity_out: int = Body(...),
    notes: str | None = Body(None),
):
    db = SessionLocal()

    try:
        event = db.execute(select(Event).where(Event.id == event_id)).scalar_one_or_none()
        if not event:
            raise HTTPException(status_code=404, detail="Evento non trovato")

        stock = db.execute(select(StockCard).where(StockCard.id == stock_card_id)).scalar_one_or_none()
        if not stock:
            raise HTTPException(status_code=404, detail="Scheda stock non trovata")

        if quantity_out <= 0:
            raise HTTPException(status_code=400, detail="La quantità in uscita deve essere maggiore di zero")

        if quantity_out > stock.quantity:
            raise HTTPException(status_code=400, detail="Quantità stock insufficiente")

        stock.quantity -= quantity_out

        event_stock = EventStock(
            event_id=event.id,
            stock_card_id=stock.id,
            quantity_out=quantity_out,
            quantity_returned=0,
            notes=notes,
        )
        db.add(event_stock)

        movement = StockMovement(
            stock_card_id=stock.id,
            movement_type="UNLOAD",
            quantity=quantity_out,
            notes=f"Uscita per evento: {event.name}" if not notes else f"Uscita per evento: {event.name} - {notes}",
        )
        db.add(movement)

        db.commit()
        db.refresh(event_stock)
        return event_stock
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# --- EVENT ASSET RETURN/MISSING & EVENT STOCK RETURN ENDPOINTS ---

@app.post("/events/{event_id}/assets/{event_asset_id}/return")
def return_event_asset(event_id: int, event_asset_id: int):
    db = SessionLocal()

    try:
        event_asset = db.execute(
            select(EventAsset).where(
                EventAsset.id == event_asset_id,
                EventAsset.event_id == event_id,
            )
        ).scalar_one_or_none()

        if not event_asset:
            raise HTTPException(status_code=404, detail="Asset evento non trovato")

        asset = db.execute(select(Asset).where(Asset.id == event_asset.asset_id)).scalar_one_or_none()
        if not asset:
            raise HTTPException(status_code=404, detail="Asset non trovato")

        event_asset.status = "RETURNED"
        event_asset.returned_at = datetime.utcnow()
        asset.status = "IN_SEDE"

        create_asset_log(
            db,
            asset.id,
            "EVENT_RETURN",
            "Asset rientrato da evento",
        )

        db.commit()
        db.refresh(event_asset)
        return event_asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.post("/events/{event_id}/assets/{event_asset_id}/missing")
def mark_event_asset_missing(event_id: int, event_asset_id: int):
    db = SessionLocal()

    try:
        event_asset = db.execute(
            select(EventAsset).where(
                EventAsset.id == event_asset_id,
                EventAsset.event_id == event_id,
            )
        ).scalar_one_or_none()

        if not event_asset:
            raise HTTPException(status_code=404, detail="Asset evento non trovato")

        asset = db.execute(select(Asset).where(Asset.id == event_asset.asset_id)).scalar_one_or_none()
        if not asset:
            raise HTTPException(status_code=404, detail="Asset non trovato")

        event_asset.status = "MISSING"
        event_asset.returned_at = None
        asset.status = "MANCANTE"

        create_asset_log(
            db,
            asset.id,
            "EVENT_MISSING",
            "Asset segnato mancante durante un evento",
        )

        db.commit()
        db.refresh(event_asset)
        return event_asset
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.post("/events/{event_id}/stocks/{event_stock_id}/return")
def return_event_stock(
    event_id: int,
    event_stock_id: int,
    quantity_returned: int = Body(...),
    notes: str | None = Body(None),
):
    db = SessionLocal()

    try:
        event = db.execute(select(Event).where(Event.id == event_id)).scalar_one_or_none()
        if not event:
            raise HTTPException(status_code=404, detail="Evento non trovato")

        event_stock = db.execute(
            select(EventStock).where(
                EventStock.id == event_stock_id,
                EventStock.event_id == event_id,
            )
        ).scalar_one_or_none()

        if not event_stock:
            raise HTTPException(status_code=404, detail="Stock evento non trovato")

        if quantity_returned <= 0:
            raise HTTPException(status_code=400, detail="La quantità rientrata deve essere maggiore di zero")

        remaining_returnable = event_stock.quantity_out - event_stock.quantity_returned
        if quantity_returned > remaining_returnable:
            raise HTTPException(
                status_code=400,
                detail="La quantità rientrata supera quella ancora da registrare come rientrata."
            )

        stock = db.execute(select(StockCard).where(StockCard.id == event_stock.stock_card_id)).scalar_one_or_none()
        if not stock:
            raise HTTPException(status_code=404, detail="Scheda stock non trovata")

        event_stock.quantity_returned += quantity_returned
        stock.quantity += quantity_returned

        movement = StockMovement(
            stock_card_id=stock.id,
            movement_type="RETURN",
            quantity=quantity_returned,
            notes=f"Rientro da evento: {event.name}" if not notes else f"Rientro da evento: {event.name} - {notes}",
        )
        db.add(movement)

        db.commit()
        db.refresh(event_stock)
        return event_stock
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# --- EVENT CLOSE/CANCEL ENDPOINTS ---

@app.post("/events/{event_id}/close")
def close_event(event_id: int):
    db = SessionLocal()

    try:
        event = db.execute(select(Event).where(Event.id == event_id)).scalar_one_or_none()
        if not event:
            raise HTTPException(status_code=404, detail="Evento non trovato")

        if event.status == "CLOSED":
            raise HTTPException(status_code=400, detail="Evento già chiuso")

        if event.status == "CANCELLED":
            raise HTTPException(status_code=400, detail="Evento annullato: non può essere chiuso")

        open_assets = db.query(EventAsset).filter(
            EventAsset.event_id == event_id,
            EventAsset.status == "OUT",
        ).count()

        event_stocks = db.query(EventStock).filter(EventStock.event_id == event_id).all()
        open_stock_returns = sum(
            max(0, event_stock.quantity_out - event_stock.quantity_returned)
            for event_stock in event_stocks
        )

        if open_assets > 0 or open_stock_returns > 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Impossibile chiudere evento: materiale ancora da rientrare. "
                    f"Asset fuori: {open_assets}. Consumabili da rientrare: {open_stock_returns}."
                ),
            )

        event.status = "CLOSED"

        db.commit()
        db.refresh(event)
        return event
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.post("/events/{event_id}/cancel")
def cancel_event(event_id: int):
    db = SessionLocal()

    try:
        event = db.execute(select(Event).where(Event.id == event_id)).scalar_one_or_none()
        if not event:
            raise HTTPException(status_code=404, detail="Evento non trovato")

        if event.status == "CLOSED":
            raise HTTPException(status_code=400, detail="Evento già chiuso: non può essere annullato")

        if event.status == "CANCELLED":
            raise HTTPException(status_code=400, detail="Evento già annullato")

        linked_assets = db.query(EventAsset).filter(EventAsset.event_id == event_id).count()
        linked_stocks = db.query(EventStock).filter(EventStock.event_id == event_id).count()

        if linked_assets > 0 or linked_stocks > 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Impossibile annullare evento: sono già presenti materiali collegati. "
                    "Prima registra i rientri o usa la chiusura evento."
                ),
            )

        event.status = "CANCELLED"

        db.commit()
        db.refresh(event)
        return event
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
